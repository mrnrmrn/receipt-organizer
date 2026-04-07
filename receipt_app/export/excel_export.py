from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from receipt_app.config import AppConfig, DEFAULT_CONFIG
from receipt_app.models import ExportRow, UploadedReceipt
from receipt_app.utils.images import (
    image_to_png_bytes,
    open_image_from_bytes,
    resize_for_excel,
)

THIN_SIDE = Side(style="thin", color="BFBFBF")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def build_workbook_bytes(
    person_name: str,
    rows: list[ExportRow | dict],
    receipts: list[UploadedReceipt],
    template_path: str | Path | None = None,
    month: date | str | None = None,
    config: AppConfig = DEFAULT_CONFIG,
) -> bytes:
    normalized_rows = _coerce_rows(rows)
    report_month = month or date.today().replace(day=1)
    return export_receipts_to_workbook(
        rows=normalized_rows,
        receipts=receipts,
        operator_name=person_name,
        report_month_text=report_month,
        template_path=template_path,
        config=config,
    )


def export_receipts_to_workbook(
    rows: list[ExportRow],
    receipts: list[UploadedReceipt],
    operator_name: str,
    report_month_text: date | str,
    template_path: str | Path | None = None,
    config: AppConfig = DEFAULT_CONFIG,
) -> bytes:
    workbook = _create_workbook(config)
    sheet = workbook[config.sheet_name]

    _ = template_path
    sheet[config.operator_name_cell] = operator_name
    sheet[config.month_cell] = _format_report_month(report_month_text)
    total_amount = sum((row.amount for row in rows), start=0)
    sheet[config.total_amount_cell] = int(total_amount)

    _write_rows(sheet=sheet, rows=rows, config=config)
    _embed_receipt_images(sheet=sheet, receipts=receipts, config=config)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _create_workbook(config: AppConfig) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = config.sheet_name

    _configure_main_sheet(sheet, config)
    _create_guide_sheet(workbook, config)
    return workbook


def _configure_main_sheet(sheet, config: AppConfig) -> None:
    sheet.freeze_panes = f"{config.number_column}{config.table_start_row}"
    sheet.sheet_view.showGridLines = False

    column_widths = {
        "A": 12,
        "B": 8,
        "C": 16,
        "D": 18,
        "E": 14,
        "F": 34,
        "G": 24,
        "H": 24,
        "I": 24,
        "J": 24,
        "K": 24,
    }
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[2].height = 22
    sheet.row_dimensions[3].height = 22
    sheet.row_dimensions[config.table_header_row].height = 24
    for image_row in config.image_anchor_rows:
        sheet.row_dimensions[image_row].height = 130

    labels = (
        (f"{config.label_column}1", "성명"),
        (f"{config.label_column}2", "합계"),
        (f"{config.label_column}3", "기준월"),
    )
    for cell, value in labels:
        sheet[cell] = value
        sheet[cell].font = Font(bold=True)
        sheet[cell].alignment = CENTER
        sheet[cell].fill = SUBHEADER_FILL
        sheet[cell].border = TABLE_BORDER

    for cell in (
        config.operator_name_cell,
        config.total_amount_cell,
        config.month_cell,
    ):
        sheet[cell].alignment = LEFT
        sheet[cell].border = TABLE_BORDER

    sheet[config.total_amount_cell].number_format = "#,##0"

    sheet[config.image_header_cell] = "영수증 이미지"
    sheet[config.image_header_cell].font = Font(bold=True)
    sheet[config.image_header_cell].alignment = LEFT

    headers = {
        f"{config.number_column}{config.table_header_row}": "번호",
        f"{config.category_column}{config.table_header_row}": "비목",
        f"{config.subcategory_column}{config.table_header_row}": "세목",
        f"{config.amount_column}{config.table_header_row}": "금액",
        f"{config.note_column}{config.table_header_row}": "사용내역",
    }
    for cell, value in headers.items():
        sheet[cell] = value
        sheet[cell].font = Font(bold=True)
        sheet[cell].fill = HEADER_FILL
        sheet[cell].alignment = CENTER
        sheet[cell].border = TABLE_BORDER

    for row_number in range(config.table_start_row, config.table_end_row + 1):
        for column in (
            config.number_column,
            config.category_column,
            config.subcategory_column,
            config.amount_column,
            config.note_column,
        ):
            cell = sheet[f"{column}{row_number}"]
            cell.border = TABLE_BORDER
            cell.alignment = (
                LEFT
                if column
                in (
                    config.category_column,
                    config.subcategory_column,
                    config.note_column,
                )
                else CENTER
            )
        sheet[f"{config.amount_column}{row_number}"].number_format = "#,##0"


def _create_guide_sheet(workbook: Workbook, config: AppConfig) -> None:
    guide_sheet = workbook.create_sheet(title=config.guide_sheet_name)
    guide_sheet.sheet_view.showGridLines = False
    guide_sheet.column_dimensions["A"].width = 24
    guide_sheet.column_dimensions["B"].width = 24

    guide_sheet["A1"] = "비목"
    guide_sheet["B1"] = "세목"
    for cell in (guide_sheet["A1"], guide_sheet["B1"]):
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = TABLE_BORDER

    unique_rules = sorted(set(config.category_rules.values()))
    for row_index, (category, subcategory) in enumerate(unique_rules, start=2):
        guide_sheet[f"A{row_index}"] = category
        guide_sheet[f"B{row_index}"] = subcategory
        guide_sheet[f"A{row_index}"].border = TABLE_BORDER
        guide_sheet[f"B{row_index}"].border = TABLE_BORDER


def _write_rows(sheet, rows: list[ExportRow], config: AppConfig) -> None:
    for row_number in range(config.table_start_row, config.table_end_row + 1):
        sheet[f"{config.number_column}{row_number}"] = None
        sheet[f"{config.category_column}{row_number}"] = None
        sheet[f"{config.subcategory_column}{row_number}"] = None
        sheet[f"{config.amount_column}{row_number}"] = None
        sheet[f"{config.note_column}{row_number}"] = None

    for index, export_row in enumerate(rows, start=1):
        row_number = config.table_start_row + index - 1
        if row_number > config.table_end_row:
            break
        sheet[f"{config.number_column}{row_number}"] = index
        sheet[f"{config.category_column}{row_number}"] = export_row.category
        sheet[f"{config.subcategory_column}{row_number}"] = export_row.subcategory
        sheet[f"{config.amount_column}{row_number}"] = int(export_row.amount)
        sheet[f"{config.note_column}{row_number}"] = export_row.note_text


def _embed_receipt_images(
    sheet, receipts: list[UploadedReceipt], config: AppConfig
) -> None:
    if hasattr(sheet, "_images"):
        sheet._images = []

    if not receipts:
        return

    slots = [
        f"{column}{row}"
        for row in config.image_anchor_rows
        for column in config.image_anchor_columns
    ]
    for receipt, anchor in zip(receipts[: config.max_receipt_images], slots):
        image = open_image_from_bytes(receipt.image_bytes)
        resized = resize_for_excel(image, max_width=150, max_height=170)
        excel_image = XLImage(BytesIO(image_to_png_bytes(resized)))
        excel_image.anchor = anchor
        sheet.add_image(excel_image)


def _format_report_month(report_month_text: date | str) -> str:
    if isinstance(report_month_text, date):
        return report_month_text.strftime("%Y-%m")
    return str(report_month_text)


def _coerce_rows(rows: list[ExportRow | dict]) -> list[ExportRow]:
    normalized: list[ExportRow] = []
    for index, row in enumerate(rows, start=1):
        if isinstance(row, ExportRow):
            normalized.append(row)
            continue

        amount_value = row.get("amount")
        if amount_value in (None, ""):
            continue

        normalized.append(
            ExportRow(
                number=index,
                category=(row.get("category") or "기타").strip(),
                subcategory=(row.get("subcategory") or "기타").strip(),
                amount=Decimal(str(amount_value).replace(",", "").strip()),
                vendor=(row.get("vendor") or None),
                receipt_date=row.get("receipt_date"),
                notes=(row.get("notes") or None),
            )
        )

    return normalized
